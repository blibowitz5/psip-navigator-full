#!/usr/bin/env python3
"""
Automated Scoring of RAG vs General LLM Responses
Analyzes responses and provides scores based on accuracy, completeness, and specificity
"""

import json
import re
from typing import Dict, List, Any, Tuple
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ResponseScorer:
    def __init__(self):
        self.scoring_criteria = {
            "accuracy": "Is the information factually correct?",
            "completeness": "Are all relevant details included?",
            "specificity": "Is it specific to the insurance plan?",
            "source_attribution": "Does it cite source documents?",
            "clarity": "Is the response clear and understandable?",
            "honesty": "Does it admit when it doesn't know something?"
        }
    
    def extract_numerical_values(self, text: str) -> List[float]:
        """Extract numerical values from text (dollar amounts, percentages, etc.)"""
        # Find dollar amounts
        dollar_pattern = r'\$[\d,]+(?:\.\d{2})?'
        dollars = re.findall(dollar_pattern, text)
        dollar_values = []
        for dollar in dollars:
            try:
                value = float(dollar.replace('$', '').replace(',', ''))
                dollar_values.append(value)
            except:
                pass
        
        # Find percentages
        percent_pattern = r'(\d+(?:\.\d+)?)%'
        percents = re.findall(percent_pattern, text)
        percent_values = [float(p) for p in percents]
        
        # Find other numbers
        number_pattern = r'\b\d+(?:\.\d+)?\b'
        numbers = re.findall(number_pattern, text)
        other_values = []
        for num in numbers:
            try:
                value = float(num)
                if value not in dollar_values and value not in percent_values:
                    other_values.append(value)
            except:
                pass
        
        return dollar_values + percent_values + other_values
    
    def score_accuracy(self, response: str, expected: str, question: str) -> int:
        """Score accuracy (1-4) based on factual correctness"""
        if not response or response.strip() == "":
            return 1
        
        # Check for specific insurance terms and values
        insurance_terms = ['deductible', 'copay', 'copayment', 'out-of-pocket', 'premium', 'coverage']
        has_insurance_terms = any(term in response.lower() for term in insurance_terms)
        
        # Extract numerical values from both response and expected
        response_values = self.extract_numerical_values(response)
        expected_values = self.extract_numerical_values(expected)
        
        # Check if response contains expected values
        value_accuracy = 0
        if expected_values:
            for expected_val in expected_values:
                if any(abs(resp_val - expected_val) < 0.01 for resp_val in response_values):
                    value_accuracy += 1
            value_accuracy = value_accuracy / len(expected_values)
        else:
            value_accuracy = 1  # No specific values to check
        
        # Check for policy-specific language
        policy_specific = any(term in response.lower() for term in [
            'in-network', 'out-of-network', 'policy year', 'eligible health services',
            'referral', 'specialist', 'preventive care'
        ])
        
        # Score based on multiple factors
        if value_accuracy >= 0.8 and has_insurance_terms and policy_specific:
            return 4  # Excellent
        elif value_accuracy >= 0.6 and has_insurance_terms:
            return 3  # Good
        elif value_accuracy >= 0.4 or has_insurance_terms:
            return 2  # Fair
        else:
            return 1  # Poor
    
    def score_completeness(self, response: str, question: str) -> int:
        """Score completeness (1-4) based on how much information is provided"""
        if not response or response.strip() == "":
            return 1
        
        # Check response length (basic heuristic)
        word_count = len(response.split())
        
        # Check for multiple pieces of information
        sentences = response.split('.')
        non_empty_sentences = [s.strip() for s in sentences if s.strip()]
        
        # Check for specific details
        has_specifics = any(char.isdigit() for char in response)  # Contains numbers
        has_conditions = 'if' in response.lower() or 'when' in response.lower()
        has_exceptions = 'except' in response.lower() or 'however' in response.lower()
        
        if word_count >= 50 and len(non_empty_sentences) >= 3 and has_specifics:
            return 4  # Excellent
        elif word_count >= 30 and len(non_empty_sentences) >= 2:
            return 3  # Good
        elif word_count >= 15 and len(non_empty_sentences) >= 1:
            return 2  # Fair
        else:
            return 1  # Poor
    
    def score_specificity(self, response: str, question: str) -> int:
        """Score specificity (1-4) based on how specific the answer is to the insurance plan"""
        if not response or response.strip() == "":
            return 1
        
        # Check for plan-specific language
        plan_specific_terms = [
            'aetna', 'student', 'policy', 'in-network', 'out-of-network',
            'copay', 'deductible', 'out-of-pocket', 'referral'
        ]
        
        specific_count = sum(1 for term in plan_specific_terms if term in response.lower())
        
        # Check for specific values
        has_dollar_amounts = '$' in response
        has_percentages = '%' in response
        has_specific_numbers = bool(re.search(r'\b\d+\b', response))
        
        if specific_count >= 4 and (has_dollar_amounts or has_percentages):
            return 4  # Excellent
        elif specific_count >= 3 and has_specific_numbers:
            return 3  # Good
        elif specific_count >= 2:
            return 2  # Fair
        else:
            return 1  # Poor
    
    def score_source_attribution(self, response: str, contexts: List[Dict]) -> int:
        """Score source attribution (1-4) based on whether sources are cited"""
        if not response or response.strip() == "":
            return 1
        
        # Check for source citations
        has_sources = 'source' in response.lower() or 'page' in response.lower()
        has_document_refs = any(doc in response.lower() for doc in [
            'master_policy', 'summary_of_benefits', 'outline_of_coverage'
        ])
        
        # Check if contexts were provided
        has_contexts = len(contexts) > 0
        
        if has_sources and has_document_refs and has_contexts:
            return 4  # Excellent
        elif has_sources and has_contexts:
            return 3  # Good
        elif has_contexts:
            return 2  # Fair
        else:
            return 1  # Poor
    
    def score_clarity(self, response: str) -> int:
        """Score clarity (1-4) based on how clear and understandable the response is"""
        if not response or response.strip() == "":
            return 1
        
        # Check for clear structure
        has_sentences = '.' in response
        has_paragraphs = '\n' in response
        
        # Check for clear language
        unclear_terms = ['unclear', 'confusing', 'complicated', 'unable to determine']
        has_unclear_terms = any(term in response.lower() for term in unclear_terms)
        
        # Check for helpful formatting
        has_lists = any(char in response for char in ['•', '-', '*', '1.', '2.'])
        has_emphasis = any(char in response for char in ['**', '__', '##'])
        
        if has_sentences and not has_unclear_terms and (has_lists or has_emphasis):
            return 4  # Excellent
        elif has_sentences and not has_unclear_terms:
            return 3  # Good
        elif has_sentences:
            return 2  # Fair
        else:
            return 1  # Poor
    
    def score_honesty(self, response: str, question: str) -> int:
        """Score honesty (1-4) based on whether the LLM admits uncertainty when appropriate"""
        if not response or response.strip() == "":
            return 1
        
        # Check for honest uncertainty
        honest_phrases = [
            'i am unable to', 'i cannot', 'i don\'t know', 'i\'m not sure',
            'unable to answer', 'cannot determine', 'not available',
            'not found in', 'not specified', 'not mentioned'
        ]
        
        has_honest_uncertainty = any(phrase in response.lower() for phrase in honest_phrases)
        
        # Check for overconfident incorrect statements
        overconfident_phrases = [
            'definitely', 'certainly', 'always', 'never', 'all', 'every'
        ]
        has_overconfident = any(phrase in response.lower() for phrase in overconfident_phrases)
        
        # Check for hedging language (good)
        hedging_phrases = [
            'may', 'might', 'could', 'likely', 'typically', 'generally'
        ]
        has_hedging = any(phrase in response.lower() for phrase in hedging_phrases)
        
        if has_honest_uncertainty and not has_overconfident:
            return 4  # Excellent
        elif has_hedging and not has_overconfident:
            return 3  # Good
        elif not has_overconfident:
            return 2  # Fair
        else:
            return 1  # Poor
    
    def calculate_overall_score(self, scores: Dict[str, int]) -> int:
        """Calculate overall score as average of all criteria"""
        if not scores:
            return 1
        
        total_score = sum(scores.values())
        average_score = total_score / len(scores)
        
        # Round to nearest integer
        return round(average_score)
    
    def score_response(self, response: str, expected: str, question: str, 
                      category: str, contexts: List[Dict] = None) -> Dict[str, Any]:
        """Score a single response across all criteria"""
        if contexts is None:
            contexts = []
        
        scores = {
            "accuracy": self.score_accuracy(response, expected, question),
            "completeness": self.score_completeness(response, question),
            "specificity": self.score_specificity(response, question),
            "source_attribution": self.score_source_attribution(response, contexts),
            "clarity": self.score_clarity(response),
            "honesty": self.score_honesty(response, question)
        }
        
        overall_score = self.calculate_overall_score(scores)
        
        return {
            "scores": scores,
            "overall_score": overall_score,
            "criteria_explanation": self.scoring_criteria
        }
    
    def create_scored_workbook(self, results_file: str = "LLM_Evaluation_Results.xlsx", 
                             output_file: str = "LLM_Evaluation_Scored.xlsx"):
        """Create a new workbook with automated scores"""
        
        # Load the results workbook
        wb = openpyxl.load_workbook(results_file)
        ws = wb["LLM Evaluation Results"]
        
        # Create new workbook with scores
        new_wb = Workbook()
        new_wb.remove(new_wb.active)
        
        # Create main results sheet
        ws_scored = new_wb.create_sheet("Scored Results")
        
        # Headers with additional score columns
        headers = [
            "Question #", "Question", "Expected Answer", "Category", "Difficulty",
            "Nelly 1.0 Answer", "Nelly 1.0 Error", "GPT-4o-mini Answer", "GPT-4o-mini Error",
            "Nelly 1.0 Overall Score", "GPT-4o-mini Overall Score",
            "Nelly 1.0 Accuracy", "Nelly 1.0 Completeness", "Nelly 1.0 Specificity", "Nelly 1.0 Source Attr", "Nelly 1.0 Clarity", "Nelly 1.0 Honesty",
            "GPT-4o-mini Accuracy", "GPT-4o-mini Completeness", "GPT-4o-mini Specificity", "GPT-4o-mini Source Attr", "GPT-4o-mini Clarity", "GPT-4o-mini Honesty",
            "Nelly 1.0 Notes", "GPT-4o-mini Notes", "Nelly 1.0 Model", "GPT-4o-mini Model", "Context Sources"
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws_scored.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Process each row
        scorer = ResponseScorer()
        
        for row in range(2, ws.max_row + 1):
            # Get data from original sheet
            question = ws.cell(row=row, column=2).value or ""
            expected = ws.cell(row=row, column=3).value or ""
            category = ws.cell(row=row, column=4).value or ""
            rag_answer = ws.cell(row=row, column=6).value or ""
            rag_error = ws.cell(row=row, column=7).value or ""
            general_answer = ws.cell(row=row, column=8).value or ""
            general_error = ws.cell(row=row, column=9).value or ""
            contexts_str = ws.cell(row=row, column=16).value or ""
            
            # Parse contexts (simplified)
            contexts = []
            if contexts_str:
                context_parts = contexts_str.split("; ")
                for part in context_parts:
                    if "(" in part and ")" in part:
                        source = part.split(" (")[0]
                        page_info = part.split(" (")[1].rstrip(")")
                        contexts.append({
                            "metadata": {
                                "source_file": source,
                                "page_number": page_info
                            }
                        })
            
            # Score RAG response
            rag_scores = {}
            rag_overall = 1
            if rag_answer and not rag_error:
                rag_scoring = scorer.score_response(rag_answer, expected, question, category, contexts)
                rag_scores = rag_scoring["scores"]
                rag_overall = rag_scoring["overall_score"]
            
            # Score General LLM response
            general_scores = {}
            general_overall = 1
            if general_answer and not general_error:
                general_scoring = scorer.score_response(general_answer, expected, question, category, [])
                general_scores = general_scoring["scores"]
                general_overall = general_scoring["overall_score"]
            
            # Write data to new sheet
            ws_scored.cell(row=row, column=1, value=row-1)  # Question #
            ws_scored.cell(row=row, column=2, value=question)
            ws_scored.cell(row=row, column=3, value=expected)
            ws_scored.cell(row=row, column=4, value=category)
            ws_scored.cell(row=row, column=5, value=ws.cell(row=row, column=5).value)  # Difficulty
            ws_scored.cell(row=row, column=6, value=rag_answer)
            ws_scored.cell(row=row, column=7, value=rag_error)
            ws_scored.cell(row=row, column=8, value=general_answer)
            ws_scored.cell(row=row, column=9, value=general_error)
            
            # Overall scores
            ws_scored.cell(row=row, column=10, value=rag_overall)
            ws_scored.cell(row=row, column=11, value=general_overall)
            
            # Detailed RAG scores
            ws_scored.cell(row=row, column=12, value=rag_scores.get("accuracy", 1))
            ws_scored.cell(row=row, column=13, value=rag_scores.get("completeness", 1))
            ws_scored.cell(row=row, column=14, value=rag_scores.get("specificity", 1))
            ws_scored.cell(row=row, column=15, value=rag_scores.get("source_attribution", 1))
            ws_scored.cell(row=row, column=16, value=rag_scores.get("clarity", 1))
            ws_scored.cell(row=row, column=17, value=rag_scores.get("honesty", 1))
            
            # Detailed General LLM scores
            ws_scored.cell(row=row, column=18, value=general_scores.get("accuracy", 1))
            ws_scored.cell(row=row, column=19, value=general_scores.get("completeness", 1))
            ws_scored.cell(row=row, column=20, value=general_scores.get("specificity", 1))
            ws_scored.cell(row=row, column=21, value=general_scores.get("source_attribution", 1))
            ws_scored.cell(row=row, column=22, value=general_scores.get("clarity", 1))
            ws_scored.cell(row=row, column=23, value=general_scores.get("honesty", 1))
            
            # Notes and other columns
            ws_scored.cell(row=row, column=24, value="")  # RAG Notes
            ws_scored.cell(row=row, column=25, value="")  # General LLM Notes
            ws_scored.cell(row=row, column=26, value="Nelly 1.0")  # RAG Model - Updated name
            ws_scored.cell(row=row, column=27, value="GPT-4o-mini")  # General Model - Updated name
            ws_scored.cell(row=row, column=28, value=contexts_str)  # Context Sources
        
        # Auto-adjust column widths
        for column in ws_scored.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_scored.column_dimensions[column_letter].width = adjusted_width
        
        # Create summary sheet
        ws_summary = new_wb.create_sheet("Scoring Summary")
        
        summary_content = [
            "AUTOMATED SCORING SUMMARY",
            "",
            "LLM COMPARISON:",
            "• Nelly 1.0: RAG-based LLM trained on insurance documents",
            "• GPT-4o-mini: General-purpose LLM without document access",
            "",
            "SCORING CRITERIA (1-4 scale):",
            "• Accuracy: Is the information factually correct?",
            "• Completeness: Are all relevant details included?",
            "• Specificity: Is it specific to the insurance plan?",
            "• Source Attribution: Does it cite source documents?",
            "• Clarity: Is the response clear and understandable?",
            "• Honesty: Does it admit when it doesn't know something?",
            "",
            "SCORING METHODOLOGY:",
            "• Automated analysis of response content",
            "• Pattern matching for insurance-specific terms",
            "• Numerical value extraction and comparison",
            "• Source citation detection",
            "• Language clarity assessment",
            "",
            "INTERPRETATION:",
            "• 4 = Excellent: Meets all criteria well",
            "• 3 = Good: Meets most criteria adequately",
            "• 2 = Fair: Meets some criteria but has gaps",
            "• 1 = Poor: Fails to meet most criteria",
            "",
            "EXPECTED RESULTS:",
            "• Nelly 1.0 should excel at plan-specific details",
            "• GPT-4o-mini may provide more generic responses",
            "• Nelly 1.0 should have better source attribution",
            "• GPT-4o-mini may be more conversational",
            "",
            "NOTE: These are automated scores. Manual review recommended for final assessment."
        ]
        
        for i, content in enumerate(summary_content, 1):
            ws_summary[f'A{i}'] = content
            if i == 1:  # Title
                ws_summary[f'A{i}'].font = Font(bold=True, size=14)
            elif i in [3, 9, 15]:  # Section headers
                ws_summary[f'A{i}'].font = Font(bold=True)
            ws_summary.column_dimensions['A'].width = 100
        
        # Save workbook
        new_wb.save(output_file)
        print(f"Scored evaluation results saved to: {output_file}")
        
        return output_file

def main():
    """Main scoring function"""
    print("🎯 Starting Automated Response Scoring...")
    
    scorer = ResponseScorer()
    
    # Create scored workbook
    output_file = scorer.create_scored_workbook()
    
    print(f"\n✅ Automated scoring complete!")
    print(f"📁 Scored results saved to: {output_file}")
    print(f"📊 Review the 'Scored Results' sheet for detailed scores")
    print(f"📋 Check the 'Scoring Summary' sheet for methodology")
    print(f"\n💡 Next steps:")
    print(f"   1. Review automated scores")
    print(f"   2. Adjust scores manually if needed")
    print(f"   3. Calculate category averages")
    print(f"   4. Identify strengths and weaknesses")

if __name__ == "__main__":
    main()
