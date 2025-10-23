#!/usr/bin/env python3
"""
LLM Evaluation Runner
Compares performance of our RAG-based LLM vs general LLM on evaluation questions
"""

import json
import requests
import time
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
from typing import Dict, List, Any

class LLMEvaluator:
    def __init__(self, backend_url: str = "http://localhost:8000"):
        self.backend_url = backend_url
        self.openai_api_key = os.environ.get("OPENAI_API_KEY")
        
    def test_rag_llm(self, question: str, n_context: int = 5) -> Dict[str, Any]:
        """Test our RAG-based LLM"""
        try:
            response = requests.post(
                f"{self.backend_url}/ask",
                json={"question": question, "n_context": n_context},
                timeout=30
            )
            if response.status_code == 200:
                return response.json()
            else:
                return {"error": f"HTTP {response.status_code}: {response.text}"}
        except Exception as e:
            return {"error": f"Request failed: {str(e)}"}
    
    def test_general_llm(self, question: str) -> Dict[str, Any]:
        """Test general LLM without RAG context"""
        try:
            import openai
            openai.api_key = self.openai_api_key
            
            response = openai.ChatCompletion.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "You are a helpful assistant answering questions about health insurance. Be honest if you don't have specific information."},
                    {"role": "user", "content": question}
                ],
                temperature=0.2,
                max_tokens=500
            )
            
            return {
                "answer": response.choices[0].message.content,
                "model": "gpt-4o-mini"
            }
        except Exception as e:
            return {"error": f"OpenAI call failed: {str(e)}"}
    
    def evaluate_question(self, qa_item: Dict[str, Any]) -> Dict[str, Any]:
        """Evaluate a single question with both LLMs"""
        question = qa_item["question"]
        expected_answer = qa_item.get("answer", "")
        category = qa_item.get("category", "")
        difficulty = qa_item.get("difficulty", "")
        
        print(f"Evaluating: {question[:60]}...")
        
        # Test RAG LLM
        rag_result = self.test_rag_llm(question)
        rag_answer = rag_result.get("answer", "No answer provided")
        rag_error = rag_result.get("error", "")
        
        # Test General LLM
        general_result = self.test_general_llm(question)
        general_answer = general_result.get("answer", "No answer provided")
        general_error = general_result.get("error", "")
        
        return {
            "question": question,
            "expected_answer": expected_answer,
            "category": category,
            "difficulty": difficulty,
            "rag_answer": rag_answer,
            "rag_error": rag_error,
            "general_answer": general_answer,
            "general_error": general_error,
            "rag_contexts": rag_result.get("contexts", []),
            "rag_model": rag_result.get("model", "RAG-based"),
            "general_model": general_result.get("model", "gpt-4o-mini")
        }
    
    def run_evaluation(self, dataset_path: str = "evaluation_qa_dataset.json") -> List[Dict[str, Any]]:
        """Run evaluation on all questions"""
        print("Loading evaluation dataset...")
        with open(dataset_path, 'r') as f:
            dataset = json.load(f)
        
        questions = dataset["questions"]
        results = []
        
        print(f"Evaluating {len(questions)} questions...")
        
        for i, qa_item in enumerate(questions, 1):
            print(f"\n--- Question {i}/{len(questions)} ---")
            result = self.evaluate_question(qa_item)
            results.append(result)
            
            # Add delay to avoid rate limiting
            time.sleep(1)
        
        return results
    
    def create_evaluation_workbook(self, results: List[Dict[str, Any]], output_path: str = "LLM_Evaluation_Results.xlsx"):
        """Create Excel workbook with evaluation results"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create main results sheet
        ws_results = wb.create_sheet("LLM Evaluation Results")
        
        # Headers
        headers = [
            "Question #", "Question", "Expected Answer", "Category", "Difficulty",
            "RAG Answer", "RAG Error", "General LLM Answer", "General LLM Error",
            "RAG Score (1-4)", "General LLM Score (1-4)", "RAG Notes", "General LLM Notes",
            "RAG Model", "General Model", "Context Sources"
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws_results.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row, result in enumerate(results, 2):
            ws_results.cell(row=row, column=1, value=row-1)  # Question #
            ws_results.cell(row=row, column=2, value=result["question"])
            ws_results.cell(row=row, column=3, value=result["expected_answer"])
            ws_results.cell(row=row, column=4, value=result["category"])
            ws_results.cell(row=row, column=5, value=result["difficulty"])
            ws_results.cell(row=row, column=6, value=result["rag_answer"])
            ws_results.cell(row=row, column=7, value=result["rag_error"])
            ws_results.cell(row=row, column=8, value=result["general_answer"])
            ws_results.cell(row=row, column=9, value=result["general_error"])
            ws_results.cell(row=row, column=10, value="")  # RAG Score - to be filled manually
            ws_results.cell(row=row, column=11, value="")  # General LLM Score - to be filled manually
            ws_results.cell(row=row, column=12, value="")  # RAG Notes - to be filled manually
            ws_results.cell(row=row, column=13, value="")  # General LLM Notes - to be filled manually
            ws_results.cell(row=row, column=14, value=result["rag_model"])
            ws_results.cell(row=row, column=15, value=result["general_model"])
            
            # Context sources
            contexts = result.get("rag_contexts", [])
            sources = []
            for ctx in contexts:
                metadata = ctx.get("metadata", {})
                source = metadata.get("source_file", "Unknown")
                page = metadata.get("page_number", "Unknown")
                sources.append(f"{source} (page {page})")
            ws_results.cell(row=row, column=16, value="; ".join(sources))
        
        # Auto-adjust column widths
        for column in ws_results.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_results.column_dimensions[column_letter].width = adjusted_width
        
        # Create summary sheet
        ws_summary = wb.create_sheet("Summary & Instructions")
        
        summary_content = [
            "LLM EVALUATION RESULTS - INSTRUCTIONS",
            "",
            "HOW TO USE THIS WORKBOOK:",
            "1. Review each question in the 'LLM Evaluation Results' sheet",
            "2. For each LLM (RAG vs General):",
            "   - Score the response: 1=Poor, 2=Fair, 3=Good, 4=Excellent",
            "   - Add notes about accuracy, completeness, or issues",
            "3. Use the 'Summary' sheet to track overall performance",
            "",
            "SCORING GUIDE:",
            "1 = Poor: Incorrect information, misleading, or unhelpful",
            "2 = Fair: Partially correct but missing important details",
            "3 = Good: Mostly correct with minor gaps or issues",
            "4 = Excellent: Accurate, complete, and well-explained",
            "",
            "EVALUATION CRITERIA:",
            "- Accuracy: Is the information correct?",
            "- Completeness: Are all relevant details included?",
            "- Clarity: Is the response clear and understandable?",
            "- Source Attribution: Does it reference specific documents?",
            "- Honesty: Does the LLM admit when it doesn't know something?",
            "",
            "COMPARISON FOCUS:",
            "- RAG LLM: Uses your specific insurance documents",
            "- General LLM: Uses general knowledge only",
            "- Which provides more accurate, specific answers?",
            "- Which handles edge cases better?",
            "",
            f"TOTAL QUESTIONS EVALUATED: {len(results)}",
            "",
            "CATEGORY BREAKDOWN:",
        ]
        
        # Add category breakdown
        categories = {}
        for result in results:
            cat = result["category"]
            if cat not in categories:
                categories[cat] = 0
            categories[cat] += 1
        
        for cat, count in categories.items():
            summary_content.append(f"{cat.replace('_', ' ').title()}: {count} questions")
        
        # Write summary content
        for i, content in enumerate(summary_content, 1):
            ws_summary[f'A{i}'] = content
            if i == 1:  # Title
                ws_summary[f'A{i}'].font = Font(bold=True, size=14)
            elif i in [3, 9, 15, 21, 27]:  # Section headers
                ws_summary[f'A{i}'].font = Font(bold=True)
            ws_summary.column_dimensions['A'].width = 100
        
        # Save workbook
        wb.save(output_path)
        print(f"\nEvaluation results saved to: {output_path}")
        
        return output_path

def main():
    """Main evaluation function"""
    print("🚀 Starting LLM Evaluation...")
    
    # Check if backend is running
    try:
        response = requests.get("http://localhost:8000/health", timeout=5)
        if response.status_code != 200:
            print("❌ Backend not running. Please start the backend first:")
            print("   python3 -m uvicorn backend_api:app --host 0.0.0.0 --port 8000")
            return
    except:
        print("❌ Backend not running. Please start the backend first:")
        print("   python3 -m uvicorn backend_api:app --host 0.0.0.0 --port 8000")
        return
    
    # Check OpenAI API key
    if not os.environ.get("OPENAI_API_KEY"):
        print("❌ OPENAI_API_KEY not set. Please set it first:")
        print("   export OPENAI_API_KEY=your_key_here")
        return
    
    # Initialize evaluator
    evaluator = LLMEvaluator()
    
    # Run evaluation
    print("📊 Running evaluation on all questions...")
    results = evaluator.run_evaluation()
    
    # Create workbook
    print("📝 Creating evaluation workbook...")
    output_path = evaluator.create_evaluation_workbook(results)
    
    print(f"\n✅ Evaluation complete!")
    print(f"📁 Results saved to: {output_path}")
    print(f"📊 Evaluated {len(results)} questions")
    print(f"🔍 Compare RAG vs General LLM performance in the Excel file")

if __name__ == "__main__":
    main()
