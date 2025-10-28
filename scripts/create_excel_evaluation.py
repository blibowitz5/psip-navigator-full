import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

class ExcelEvaluationGenerator:
    def __init__(self):
        self.workbook = Workbook()
        self.setup_styles()
    
    def setup_styles(self):
        """Setup Excel styling"""
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.question_font = Font(bold=True)
        self.answer_font = Font()
        self.category_font = Font(italic=True, color="666666")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    def load_evaluation_data(self, json_file="evaluation_qa_dataset.json"):
        """Load the evaluation data from JSON"""
        with open(json_file, 'r') as f:
            return json.load(f)
    
    def create_main_evaluation_sheet(self, data):
        """Create the main evaluation sheet"""
        ws = self.workbook.active
        ws.title = "LLM Evaluation Results"
        
        # Prepare data for DataFrame
        rows = []
        for i, qa in enumerate(data['questions'], 1):
            row = {
                'Question #': i,
                'Category': qa['category'].replace('_', ' ').title(),
                'Difficulty': qa['difficulty'].title(),
                'Question': qa['question'],
                'Correct Answer': qa['answer'],
                'Source Files': ', '.join(qa['source_files']),
                'Can Answer': 'Yes' if 'note' not in qa else 'No',
                'Notes': qa.get('note', ''),
                'LLM 1 Score': '',  # Empty for manual entry
                'LLM 1 Response': '',  # Empty for manual entry
                'LLM 1 Notes': '',  # Empty for manual entry
                'LLM 2 Score': '',  # Empty for manual entry
                'LLM 2 Response': '',  # Empty for manual entry
                'LLM 2 Notes': '',  # Empty for manual entry
                'LLM 3 Score': '',  # Empty for manual entry
                'LLM 3 Response': '',  # Empty for manual entry
                'LLM 3 Notes': '',  # Empty for manual entry
            }
            rows.append(row)
        
        # Convert to DataFrame
        df = pd.DataFrame(rows)
        
        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Style the worksheet
        self.style_evaluation_sheet(ws, len(data['questions']))
        
        return ws
    
    def style_evaluation_sheet(self, ws, num_questions):
        """Apply styling to the evaluation sheet"""
        # Header row styling
        for cell in ws[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Column widths
        column_widths = {
            'A': 12,  # Question #
            'B': 20,  # Category
            'C': 12,  # Difficulty
            'D': 60,  # Question
            'E': 80,  # Correct Answer
            'F': 30,  # Source Files
            'G': 12,  # Can Answer
            'H': 30,  # Notes
            'I': 12,  # LLM 1 Score
            'J': 60,  # LLM 1 Response
            'K': 30,  # LLM 1 Notes
            'L': 12,  # LLM 2 Score
            'M': 60,  # LLM 2 Response
            'N': 30,  # LLM 2 Notes
            'O': 12,  # LLM 3 Score
            'P': 60,  # LLM 3 Response
            'Q': 30,  # LLM 3 Notes
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Apply styling to data rows
        for row in range(2, num_questions + 2):
            # Question number
            ws[f'A{row}'].font = self.question_font
            ws[f'A{row}'].alignment = self.center_alignment
            
            # Category
            ws[f'B{row}'].font = self.category_font
            
            # Difficulty
            ws[f'C{row}'].alignment = self.center_alignment
            
            # Question
            ws[f'D{row}'].font = self.question_font
            ws[f'D{row}'].alignment = self.wrap_alignment
            
            # Answer
            ws[f'E{row}'].alignment = self.wrap_alignment
            
            # Source files
            ws[f'F{row}'].alignment = self.wrap_alignment
            
            # Can Answer
            ws[f'G{row}'].alignment = self.center_alignment
            
            # Notes
            ws[f'H{row}'].alignment = self.wrap_alignment
            
            # LLM response columns
            for col in ['J', 'M', 'P']:  # Response columns
                ws[f'{col}{row}'].alignment = self.wrap_alignment
            
            # LLM notes columns
            for col in ['K', 'N', 'Q']:  # Notes columns
                ws[f'{col}{row}'].alignment = self.wrap_alignment
            
            # Score columns
            for col in ['I', 'L', 'O']:  # Score columns
                ws[f'{col}{row}'].alignment = self.center_alignment
            
            # Apply borders to all cells
            for col in range(1, 18):  # A to Q
                ws.cell(row=row, column=col).border = self.border
        
        # Freeze panes
        ws.freeze_panes = 'A2'
    
    def create_summary_sheet(self, data):
        """Create a summary sheet with statistics"""
        ws = self.workbook.create_sheet("Summary & Instructions")
        
        # Instructions
        instructions = [
            "LLM EVALUATION WORKBOOK - INSTRUCTIONS",
            "",
            "HOW TO USE THIS WORKBOOK:",
            "1. Review each question in the 'LLM Evaluation Results' sheet",
            "2. For each LLM you're testing:",
            "   - Enter the LLM's response in the 'Response' column",
            "   - Score the response: 1=Poor, 2=Fair, 3=Good, 4=Excellent",
            "   - Add notes about accuracy, completeness, or issues",
            "3. Use the 'Summary' sheet to track overall performance",
            "4. Questions marked 'Can Answer: No' should be handled honestly by LLMs",
            "",
            "SCORING GUIDE:",
            "1 = Poor: Incorrect information, misleading, or unhelpful",
            "2 = Fair: Partially correct but missing important details",
            "3 = Good: Mostly correct with minor gaps or issues",
            "4 = Excellent: Accurate, complete, and well-explained",
            "",
            "EVALUATION CRITERIA:",
            "- Accuracy: Is the information correct?",
            "- Completeness: Are all relevant details included?",
            "- Clarity: Is the response clear and understandable?",
            "- Honesty: Does the LLM admit when it doesn't know something?",
            "- Source Attribution: Does it reference specific documents?",
            "",
            "DATASET STATISTICS:",
            f"Total Questions: {data['metadata']['total_questions']}",
            f"Answerable Questions: {len([q for q in data['questions'] if 'note' not in q])}",
            f"Unanswerable Questions: {len([q for q in data['questions'] if 'note' in q])}",
            f"Categories: {', '.join(data['metadata']['categories'])}",
            f"Difficulty Levels: {', '.join(data['metadata']['difficulty_levels'])}",
            "",
            "CATEGORY BREAKDOWN:",
        ]
        
        # Add category breakdown
        categories = {}
        for qa in data['questions']:
            cat = qa['category']
            if cat not in categories:
                categories[cat] = {'total': 0, 'answerable': 0}
            categories[cat]['total'] += 1
            if 'note' not in qa:
                categories[cat]['answerable'] += 1
        
        for cat, stats in categories.items():
            instructions.append(f"{cat.replace('_', ' ').title()}: {stats['answerable']}/{stats['total']} answerable")
        
        # Write instructions to sheet
        for i, instruction in enumerate(instructions, 1):
            ws[f'A{i}'] = instruction
            if i == 1:  # Title
                ws[f'A{i}'].font = Font(bold=True, size=16)
            elif instruction and instruction.startswith(('HOW TO', 'SCORING', 'EVALUATION', 'DATASET', 'CATEGORY')):
                ws[f'A{i}'].font = Font(bold=True)
        
        # Set column width
        ws.column_dimensions['A'].width = 80
        
        return ws
    
    def create_performance_summary_sheet(self, data):
        """Create a performance summary sheet"""
        ws = self.workbook.create_sheet("Performance Summary")
        
        # Headers
        headers = [
            "LLM Name", "Total Questions", "Average Score", "Accuracy Rate", 
            "Easy Questions Avg", "Medium Questions Avg", "Hard Questions Avg",
            "Answerable Questions Avg", "Unanswerable Questions Avg",
            "Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Add sample rows for 3 LLMs
        sample_data = [
            ["LLM 1", "", "", "", "", "", "", "", "", "Enter LLM name and results"],
            ["LLM 2", "", "", "", "", "", "", "", "", "Enter LLM name and results"],
            ["LLM 3", "", "", "", "", "", "", "", "", "Enter LLM name and results"],
        ]
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
                if col_idx in [2, 3, 4, 5, 6, 7, 8, 9]:  # Numeric columns
                    cell.alignment = self.center_alignment
        
        # Set column widths
        column_widths = [20, 15, 15, 15, 20, 20, 20, 25, 25, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        return ws
    
    def create_category_breakdown_sheet(self, data):
        """Create a category breakdown sheet"""
        ws = self.workbook.create_sheet("Category Breakdown")
        
        # Headers
        headers = ["Category", "Total Questions", "Answerable", "Unanswerable", "Easy", "Medium", "Hard"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Calculate category statistics
        categories = {}
        for qa in data['questions']:
            cat = qa['category']
            if cat not in categories:
                categories[cat] = {
                    'total': 0, 'answerable': 0, 'unanswerable': 0,
                    'easy': 0, 'medium': 0, 'hard': 0
                }
            
            categories[cat]['total'] += 1
            if 'note' in qa:
                categories[cat]['unanswerable'] += 1
            else:
                categories[cat]['answerable'] += 1
            
            categories[cat][qa['difficulty']] += 1
        
        # Add data rows
        for row_idx, (cat, stats) in enumerate(categories.items(), 2):
            row_data = [
                cat.replace('_', ' ').title(),
                stats['total'],
                stats['answerable'],
                stats['unanswerable'],
                stats['easy'],
                stats['medium'],
                stats['hard']
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
                if col_idx > 1:  # Numeric columns
                    cell.alignment = self.center_alignment
        
        # Set column widths
        column_widths = [25, 15, 12, 12, 8, 10, 8]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        return ws
    
    def generate_excel_file(self, json_file="evaluation_qa_dataset.json", output_file="LLM_Evaluation_Workbook.xlsx"):
        """Generate the complete Excel workbook"""
        # Load data
        data = self.load_evaluation_data(json_file)
        
        # Create sheets
        self.create_main_evaluation_sheet(data)
        self.create_summary_sheet(data)
        self.create_performance_summary_sheet(data)
        self.create_category_breakdown_sheet(data)
        
        # Save workbook
        self.workbook.save(output_file)
        print(f"Excel workbook saved as: {output_file}")
        
        return output_file

def main():
    generator = ExcelEvaluationGenerator()
    output_file = generator.generate_excel_file()
    
    print(f"\nExcel workbook created successfully!")
    print(f"File: {output_file}")
    print(f"\nThe workbook contains 4 sheets:")
    print("1. LLM Evaluation Results - Main evaluation sheet with questions and scoring columns")
    print("2. Summary & Instructions - Instructions and dataset statistics")
    print("3. Performance Summary - Overall performance tracking")
    print("4. Category Breakdown - Statistics by category")
    print(f"\nReady for non-technical users to evaluate LLM performance!")

if __name__ == "__main__":
    main()
