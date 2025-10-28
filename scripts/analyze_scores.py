#!/usr/bin/env python3
"""
Analyze the scored evaluation results
Provides insights into RAG vs General LLM performance
"""

import openpyxl
from collections import defaultdict
import statistics

def analyze_scored_results(workbook_path: str = "LLM_Evaluation_Scored.xlsx"):
    """Analyze the scored evaluation results"""
    
    print("📊 Analyzing Scored Evaluation Results")
    print("=" * 50)
    
    # Load workbook
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["Scored Results"]
    
    # Collect data
    rag_overall_scores = []
    general_overall_scores = []
    category_scores = defaultdict(lambda: {"rag": [], "general": []})
    difficulty_scores = defaultdict(lambda: {"rag": [], "general": []})
    
    total_questions = 0
    successful_rag = 0
    successful_general = 0
    
    # Process each row
    for row in range(2, ws.max_row + 1):
        question = ws.cell(row=row, column=2).value
        if not question:
            continue
            
        total_questions += 1
        
        # Get scores
        rag_overall = ws.cell(row=row, column=10).value
        general_overall = ws.cell(row=row, column=11).value
        category = ws.cell(row=row, column=4).value
        difficulty = ws.cell(row=row, column=5).value
        rag_error = ws.cell(row=row, column=7).value
        general_error = ws.cell(row=row, column=9).value
        
        # Count successful responses
        if rag_overall and not rag_error:
            successful_rag += 1
            rag_overall_scores.append(rag_overall)
            category_scores[category]["rag"].append(rag_overall)
            difficulty_scores[difficulty]["rag"].append(rag_overall)
        
        if general_overall and not general_error:
            successful_general += 1
            general_overall_scores.append(general_overall)
            category_scores[category]["general"].append(general_overall)
            difficulty_scores[difficulty]["general"].append(general_overall)
    
    # Calculate overall statistics
    print(f"📈 Overall Performance:")
    print(f"   Total Questions: {total_questions}")
    print(f"   Successful Nelly 1.0 Responses: {successful_rag}")
    print(f"   Successful GPT-4o-mini Responses: {successful_general}")
    
    if rag_overall_scores:
        rag_avg = statistics.mean(rag_overall_scores)
        rag_median = statistics.median(rag_overall_scores)
        print(f"   Nelly 1.0 Average Score: {rag_avg:.2f}")
        print(f"   Nelly 1.0 Median Score: {rag_median:.2f}")
    else:
        print(f"   Nelly 1.0 Average Score: N/A (no successful responses)")
    
    if general_overall_scores:
        general_avg = statistics.mean(general_overall_scores)
        general_median = statistics.median(general_overall_scores)
        print(f"   GPT-4o-mini Average Score: {general_avg:.2f}")
        print(f"   GPT-4o-mini Median Score: {general_median:.2f}")
    else:
        print(f"   GPT-4o-mini Average Score: N/A (no successful responses)")
    
    # Category analysis
    print(f"\n📂 Performance by Category:")
    for category, scores in category_scores.items():
        if not category:
            continue
            
        rag_scores = scores["rag"]
        general_scores = scores["general"]
        
        print(f"\n   {category.replace('_', ' ').title()}:")
        if rag_scores:
            rag_avg = statistics.mean(rag_scores)
            print(f"     Nelly 1.0: {rag_avg:.2f} (n={len(rag_scores)})")
        else:
            print(f"     Nelly 1.0: N/A (no responses)")
            
        if general_scores:
            general_avg = statistics.mean(general_scores)
            print(f"     GPT-4o-mini: {general_avg:.2f} (n={len(general_scores)})")
        else:
            print(f"     GPT-4o-mini: N/A (no responses)")
    
    # Difficulty analysis
    print(f"\n🎯 Performance by Difficulty:")
    for difficulty, scores in difficulty_scores.items():
        if not difficulty:
            continue
            
        rag_scores = scores["rag"]
        general_scores = scores["general"]
        
        print(f"\n   {difficulty.title()}:")
        if rag_scores:
            rag_avg = statistics.mean(rag_scores)
            print(f"     Nelly 1.0: {rag_avg:.2f} (n={len(rag_scores)})")
        else:
            print(f"     Nelly 1.0: N/A (no responses)")
            
        if general_scores:
            general_avg = statistics.mean(general_scores)
            print(f"     GPT-4o-mini: {general_avg:.2f} (n={len(general_scores)})")
        else:
            print(f"     GPT-4o-mini: N/A (no responses)")
    
    # Key insights
    print(f"\n💡 Key Insights:")
    
    if rag_overall_scores and general_overall_scores:
        if statistics.mean(rag_overall_scores) > statistics.mean(general_overall_scores):
            print(f"   ✅ Nelly 1.0 outperforms GPT-4o-mini overall")
        elif statistics.mean(general_overall_scores) > statistics.mean(rag_overall_scores):
            print(f"   ✅ GPT-4o-mini outperforms Nelly 1.0 overall")
        else:
            print(f"   ⚖️ Nelly 1.0 and GPT-4o-mini perform similarly overall")
    
    # Success rate analysis
    rag_success_rate = (successful_rag / total_questions) * 100 if total_questions > 0 else 0
    general_success_rate = (successful_general / total_questions) * 100 if total_questions > 0 else 0
    
    print(f"   📊 Success Rates:")
    print(f"     Nelly 1.0: {rag_success_rate:.1f}% ({successful_rag}/{total_questions})")
    print(f"     GPT-4o-mini: {general_success_rate:.1f}% ({successful_general}/{total_questions})")
    
    # Category strengths
    print(f"\n🏆 Category Strengths:")
    for category, scores in category_scores.items():
        if not category or not scores["rag"] or not scores["general"]:
            continue
            
        rag_avg = statistics.mean(scores["rag"])
        general_avg = statistics.mean(scores["general"])
        
        if rag_avg > general_avg + 0.5:
            print(f"   🎯 Nelly 1.0 excels at: {category.replace('_', ' ').title()}")
        elif general_avg > rag_avg + 0.5:
            print(f"   🎯 GPT-4o-mini excels at: {category.replace('_', ' ').title()}")
    
    print(f"\n📋 Recommendations:")
    print(f"   1. Review individual responses for manual score adjustments")
    print(f"   2. Focus on categories where one approach significantly outperforms")
    print(f"   3. Consider hybrid approaches for best results")
    print(f"   4. Use insights to improve the RAG system")
    
    return {
        "total_questions": total_questions,
        "rag_scores": rag_overall_scores,
        "general_scores": general_overall_scores,
        "category_scores": dict(category_scores),
        "difficulty_scores": dict(difficulty_scores)
    }

def main():
    """Main analysis function"""
    try:
        results = analyze_scored_results()
        print(f"\n✅ Analysis complete!")
    except FileNotFoundError:
        print("❌ Scored workbook not found. Please run score_responses.py first.")
    except Exception as e:
        print(f"❌ Error analyzing results: {e}")

if __name__ == "__main__":
    main()
